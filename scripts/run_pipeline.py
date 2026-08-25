#!/usr/bin/env python3
"""Two-stage local evidence preparation and deterministic XLSX rendering."""
from __future__ import annotations
import argparse, hashlib, json, os, re, shutil, subprocess, sys, tempfile, zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("缺少 openpyxl；请安装 requirements.txt。") from exc

SUPPORTED={".docx",".pdf",".xlsx",".png",".jpg",".jpeg",".tif",".tiff",".bmp",".txt",".md"}
PERSPECTIVES={"worker":"劳动者","employer":"用人单位","neutral":"中立"}
STATUSES={"完整覆盖","部分覆盖","缺失","相互冲突","待确认"}
W_NS="http:"+"//schemas.openxmlformats.org/wordprocessingml/2006/main"
MARKER=".labor-evidence-audit-workspace"

class PipelineError(RuntimeError): pass

def stdin_json():
    try: return json.load(sys.stdin)
    except json.JSONDecodeError as exc: raise PipelineError(f"标准输入不是有效 JSON：{exc}") from exc

def digest(path: Path):
    h=hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda:f.read(1024*1024),b""): h.update(chunk)
    return h.hexdigest()

def reliable(text: str): return len(re.sub(r"\s+","",text))>=8 and "�" not in text

def convert(path: Path):
    try:
        from markitdown import MarkItDown
        md=MarkItDown(); method=getattr(md,"convert_local",md.convert)
        result=method(str(path)); text=getattr(result,"markdown",None) or getattr(result,"text_content",None) or ""
        return text, "MarkItDown 本地转换成功" if reliable(text) else "MarkItDown 未取得可靠文本"
    except ImportError: return "", "未安装 MarkItDown，已使用格式专用本地提取回退"
    except Exception as exc: return "", f"MarkItDown 转换失败：{type(exc).__name__}"

def locations(path: Path):
    suffix=path.suffix.lower(); out=[]
    if suffix==".docx":
        with zipfile.ZipFile(path) as z:
            root=ET.fromstring(z.read("word/document.xml"))
            for i,p in enumerate(root.findall(f".//{{{W_NS}}}p"),1):
                text="".join(x.text or "" for x in p.findall(f".//{{{W_NS}}}t")).strip()
                if text: out.append({"location":f"段落 {i}","text":text})
    elif suffix==".pdf":
        try:
            from pypdf import PdfReader
            for i,p in enumerate(PdfReader(str(path)).pages,1):
                text=(p.extract_text() or "").strip()
                if text: out.append({"location":f"第 {i} 页","text":text})
        except Exception: pass
    elif suffix==".xlsx":
        book=load_workbook(path,read_only=True,data_only=True)
        try:
            for sheet in book.worksheets:
                for i,row in enumerate(sheet.iter_rows(values_only=True),1):
                    vals=[str(v).strip() for v in row if v is not None and str(v).strip()]
                    if vals: out.append({"location":f"工作表“{sheet.title}”第 {i} 行","text":" | ".join(vals)})
        finally: book.close()
    return out

def ocr(path: Path):
    commands=[["tesseract",str(path),"stdout","-l","chi_sim+eng"]]
    if sys.platform=="darwin": commands.append(["swift",str(Path(__file__).with_name("ocr.swift")),str(path)])
    for cmd in commands:
        try: p=subprocess.run(cmd,capture_output=True,text=True,timeout=240,check=False)
        except (FileNotFoundError,subprocess.SubprocessError): continue
        if p.returncode==0 and reliable(p.stdout): return p.stdout.strip(), f"已使用本地 {cmd[0]} OCR；必须回看原件"
    return "", "本地 OCR 不可用或未取得可靠文本"

def extract(path: Path, workspace: Path):
    suffix=path.suffix.lower()
    text,note=((path.read_text(encoding="utf-8",errors="replace"),"文本文件本地读取") if suffix in {".txt",".md"} else convert(path))
    chunks=locations(path)
    if suffix in {".png",".jpg",".jpeg",".tif",".tiff",".bmp"} or (suffix==".pdf" and not chunks):
        ocr_text,ocr_note=ocr(path); note=f"{note}；{ocr_note}"
        if reliable(ocr_text): chunks=[{"location":"OCR 全文（页码待人工确认）","text":ocr_text}]
    if not chunks and reliable(text): chunks=[{"location":"提取文本（页码待人工确认）","text":text.strip()}]
    if not chunks: raise PipelineError(f"无法取得可靠文本：{path.name}（{note}）")
    name=f"evidence-{hashlib.sha256(str(path).encode()).hexdigest()[:12]}.md"
    md=workspace/name
    md.write_text("# 证据文件："+path.name+"\n\n"+"\n\n".join(f"## {x['location']}\n\n{x['text']}" for x in chunks)+"\n",encoding="utf-8")
    os.chmod(md,0o600)
    return {"source_name":path.name,"source_path":str(path.resolve()),"source_sha256":digest(path),"format":suffix,"markdown_path":str(md.resolve()),"locators":[x["location"] for x in chunks],"processing_note":note}

def cleanup(workspace: Path):
    if not (workspace/ MARKER).is_file(): raise PipelineError("拒绝清理：工作目录缺少安全标记")
    if workspace.resolve() in {Path("/").resolve(),Path.home().resolve()} or len(workspace.resolve().parts)<4: raise PipelineError("拒绝清理过于宽泛的目录")
    shutil.rmtree(workspace.resolve())

def prepare(args):
    if not args.privacy_confirmed: raise PipelineError("必须先取得用户对读取原始证据的明确知情确认")
    facts=stdin_json()
    if not isinstance(facts,dict) or not str(facts.get("summary","")).strip(): raise PipelineError("facts-json 必须包含非空 summary")
    paths=[Path(x) for x in args.evidence]
    if not paths: raise PipelineError("至少需要一个证据文件")
    for p in paths:
        if p.suffix.lower() not in SUPPORTED: raise PipelineError(f"不支持的证据格式：{p.suffix or '(无扩展名)'}")
        if not p.is_file(): raise PipelineError(f"证据文件不存在：{p}")
    workspace=Path(args.workspace).resolve()
    if workspace.exists() and any(workspace.iterdir()): raise PipelineError("工作目录必须不存在或为空")
    workspace.mkdir(parents=True,mode=0o700,exist_ok=True); os.chmod(workspace,0o700); (workspace/MARKER).write_text("temporary\n")
    try:
        files=[extract(p,workspace) for p in paths]
        bundle={"schema_version":"1.0","created_at":datetime.now().astimezone().isoformat(timespec="seconds"),"processing_environment":args.processing_environment,"perspective":args.perspective,"perspective_label":PERSPECTIVES[args.perspective],"facts":facts,"evidence_files":files,"analysis_instructions":{"matrix_statuses":sorted(STATUSES),"source_locator_required":True,"fact_date_controls_law_version":True,"no_authenticity_conclusion":True}}
        out=workspace/"case-bundle.json"; out.write_text(json.dumps(bundle,ensure_ascii=False,indent=2),encoding="utf-8"); os.chmod(out,0o600)
    except Exception:
        cleanup(workspace); raise
    print(out); return 0

COLUMNS={"案件事实与请求":[("类型","type"),("事项","item"),("用户陈述","statement"),("材料印证情况","support"),("差异或待确认问题","issue")],"证据清单":[("证据编号","evidence_id"),("证据名称","name"),("证据类型","type"),("来源文件","source_file"),("页码或行号","locator"),("形成时间","created_at"),("内容摘要","summary"),("真实性风险","authenticity_risk"),("合法性风险","legality_risk"),("关联性或证明力风险","relevance_risk")],"五层矩阵":[("请求事项","claim"),("法律构成要件及举证责任","element_and_burden"),("待证事实","fact_to_prove"),("对应证据及来源定位","evidence_and_locator"),("覆盖状态","status"),("分析理由","reason"),("补证动作","action")],"矛盾与风险":[("编号","id"),("冲突类型","type"),("涉及事项","subject"),("证据或陈述A","side_a"),("证据或陈述B","side_b"),("来源定位","locator"),("风险说明","risk"),("建议核实方式","action")],"缺口与补证":[("编号","id"),("请求事项","claim"),("缺失事实","missing_fact"),("建议补充材料","suggested_evidence"),("可能持有人","holder"),("建议取得方式","method"),("优先级","priority")],"法律依据":[("法律名称","title"),("条款","article"),("适用版本","version"),("效力期间","effective_period"),("相关法律事实日期","fact_date"),("官方来源","source_name"),("来源链接","source_url"),("检索时间","checked_at"),("适用说明","application")],"处理记录":[("文件或步骤","subject"),("处理状态","status"),("处理方式","method"),("定位范围","locator_scope"),("警告或核验边界","note")]} 

def validate_analysis(value):
    if not isinstance(value,dict): raise PipelineError("analysis-json 根节点必须是对象")
    result={k:value.get(k,[]) for k in COLUMNS}
    for k,v in result.items():
        if not isinstance(v,list) or any(not isinstance(x,dict) for x in v): raise PipelineError(f"analysis-json.{k} 必须是对象数组")
    for i,row in enumerate(result["五层矩阵"],1):
        if row.get("status") not in STATUSES: raise PipelineError(f"五层矩阵第 {i} 行覆盖状态无效")
        for key in ("claim","element_and_burden","fact_to_prove","status","reason"):
            if not str(row.get(key,"")).strip(): raise PipelineError(f"五层矩阵第 {i} 行缺少 {key}")
    for i,row in enumerate(result["证据清单"],1):
        if not str(row.get("source_file","")).strip() or not str(row.get("locator","")).strip(): raise PipelineError(f"证据清单第 {i} 行缺少来源文件或定位")
    result["overall_opinion"]=str(value.get("overall_opinion","未提供")); result["verification_scope"]=str(value.get("verification_scope","未说明")); return result

def style(sheet,widths):
    fill=PatternFill("solid",fgColor="1F4E78"); side=Side(style="thin",color="D9E2F3")
    for c in sheet[1]: c.fill=fill; c.font=Font(name="Microsoft YaHei",color="FFFFFF",bold=True); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for c in row: c.font=Font(name="Microsoft YaHei",size=10); c.alignment=Alignment(vertical="top",wrap_text=True); c.border=Border(left=side,right=side,top=side,bottom=side)
    for i,w in enumerate(widths,1): sheet.column_dimensions[get_column_letter(i)].width=w
    sheet.freeze_panes="A2"; sheet.auto_filter.ref=sheet.dimensions; sheet.sheet_view.showGridLines=False

def add_sheet(book,name,columns,rows):
    s=book.create_sheet(name); s.append([x[0] for x in columns])
    for row in rows: s.append(["；".join(map(str,row.get(k,[]))) if isinstance(row.get(k),list) else row.get(k,"") for _,k in columns])
    style(s,[max(12,min(48,len(title)*2+8)) for title,_ in columns])

def report(args):
    bundle_path=Path(args.case_bundle).resolve(); workspace=bundle_path.parent
    if bundle_path.name!="case-bundle.json" or not (workspace/MARKER).is_file(): raise PipelineError("case-bundle 不在受管临时工作目录中")
    output=Path(args.output).resolve()
    if output.exists(): raise PipelineError("输出文件已存在，拒绝覆盖")
    analysis=validate_analysis(stdin_json()); bundle=json.loads(bundle_path.read_text(encoding="utf-8")); matrix=analysis["五层矩阵"]; counts=Counter(x["status"] for x in matrix)
    book=Workbook(); s=book.active; s.title="汇总"; s.append(["项目","结果"])
    summary=[("案件名称",bundle["facts"].get("case_name","未命名案件")),("分析立场",bundle["perspective_label"]),("请求事项数量",len({x.get("claim") for x in matrix if x.get("claim")})),("证据数量",len(analysis["证据清单"]))]+[(x,counts[x]) for x in ("完整覆盖","部分覆盖","缺失","相互冲突","待确认")]+[("总体意见",analysis["overall_opinion"]),("法律依据核验范围",analysis["verification_scope"]),("生成时间",datetime.now().astimezone().isoformat(timespec="seconds")),("重要说明","本报告仅审计证据链覆盖和风险，不认定证据真实、合法或必然被采信，也不计算金额或期限。"),("保密提示","报告可能包含劳动争议敏感信息，请限制访问并妥善保管。")]
    for row in summary:s.append(list(row))
    style(s,[24,90])
    for name,columns in COLUMNS.items():
        rows=analysis[name]
        if name=="处理记录": rows=[{"subject":x["source_name"],"status":"成功转换","method":x["processing_note"],"locator_scope":"、".join(x["locators"][:8]),"note":"完整原文未写入报告"} for x in bundle["evidence_files"]]+rows
        add_sheet(book,name,columns,rows)
    output.parent.mkdir(parents=True,exist_ok=True); fd,tmp=tempfile.mkstemp(prefix=".labor-audit-",suffix=".xlsx",dir=output.parent); os.close(fd)
    try: book.save(tmp); Path(tmp).replace(output)
    finally: Path(tmp).unlink(missing_ok=True); book.close()
    if args.cleanup: cleanup(workspace)
    print(output); print("报告可能包含案件敏感信息，请限制访问并妥善保管。")
    return 0

def main():
    p=argparse.ArgumentParser(); sub=p.add_subparsers(dest="command",required=True)
    a=sub.add_parser("prepare"); a.add_argument("--facts-json",choices=["-"],required=True); a.add_argument("--perspective",choices=sorted(PERSPECTIVES),required=True); a.add_argument("--evidence",nargs="+",required=True); a.add_argument("--workspace",required=True); a.add_argument("--processing-environment",choices=["local","cloud"],required=True); a.add_argument("--privacy-confirmed",action="store_true")
    b=sub.add_parser("report"); b.add_argument("--case-bundle",required=True); b.add_argument("--analysis-json",choices=["-"],required=True); b.add_argument("--output",required=True); b.add_argument("--cleanup",action="store_true")
    args=p.parse_args()
    try:return prepare(args) if args.command=="prepare" else report(args)
    except (PipelineError,OSError,ValueError,zipfile.BadZipFile,ET.ParseError) as exc: print(f"拒绝处理：{exc}",file=sys.stderr); return 2
if __name__=="__main__": raise SystemExit(main())
