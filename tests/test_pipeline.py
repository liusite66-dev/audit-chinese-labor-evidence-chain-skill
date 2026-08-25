from __future__ import annotations
import hashlib, json, subprocess, sys, tempfile, unittest, zipfile
from pathlib import Path
from openpyxl import Workbook, load_workbook

ROOT=Path(__file__).resolve().parents[1]
RUNNER=ROOT/"scripts"/"run_pipeline.py"

class PipelineTests(unittest.TestCase):
    def setUp(self):
        self.temp=Path(tempfile.mkdtemp())
        self.evidence=self.temp/"解除通知.txt"
        self.evidence.write_text("公司于2026年1月5日向员工发送解除通知，理由为严重违纪。\n员工否认违纪。",encoding="utf-8")

    def prepare(self, confirmed=True):
        workspace=self.temp/"workspace"
        facts={"case_name":"张某违法解除争议","summary":"公司以严重违纪解除劳动合同，劳动者否认。","claims":["违法解除赔偿"]}
        args=[sys.executable,str(RUNNER),"prepare","--facts-json","-","--perspective","worker","--evidence",str(self.evidence),"--workspace",str(workspace),"--processing-environment","cloud"]
        if confirmed: args.append("--privacy-confirmed")
        return subprocess.run(args,input=json.dumps(facts,ensure_ascii=False),text=True,capture_output=True),workspace

    def test_privacy_gate_precedes_file_read(self):
        process,workspace=self.prepare(False)
        self.assertEqual(process.returncode,2)
        self.assertIn("明确知情确认",process.stderr)
        self.assertFalse(workspace.exists())

    def test_prepare_and_report_cleanup(self):
        before=hashlib.sha256(self.evidence.read_bytes()).hexdigest()
        process,workspace=self.prepare()
        self.assertEqual(process.returncode,0,process.stderr)
        bundle=workspace/"case-bundle.json"
        self.assertTrue(bundle.exists())
        analysis={
            "overall_opinion":"解除理由已有部分材料，但违纪事实与程序证据不足。",
            "verification_scope":"官方来源未接入；仅完成材料结构审计。",
            "案件事实与请求":[{"type":"请求事项","item":"违法解除赔偿","statement":"用户陈述","support":"解除通知印证解除行为","issue":"违纪事实待确认"}],
            "证据清单":[{"evidence_id":"E-001","name":"解除通知","type":"书证","source_file":"解除通知.txt","locator":"提取文本（页码待人工确认）","created_at":"2026-01-05","summary":"记载解除理由","authenticity_risk":"来源需核实","legality_risk":"无独立风险结论","relevance_risk":"可证明解除通知存在"}],
            "五层矩阵":[{"claim":"违法解除赔偿","element_and_burden":"解除理由及程序要件；按规则分配举证责任","fact_to_prove":"是否存在严重违纪事实","evidence_and_locator":"E-001，解除通知.txt，提取文本（页码待人工确认）","status":"部分覆盖","reason":"有解除理由记载但无违纪原始材料","action":"补充规章制度、违纪证据和送达记录"}],
            "矛盾与风险":[],"缺口与补证":[{"id":"G-001","claim":"违法解除赔偿","missing_fact":"严重违纪事实","suggested_evidence":"考勤、调查记录、制度签收","holder":"用人单位","method":"合法申请调取或提供","priority":"高"}],
            "法律依据":[],"处理记录":[]}
        output=self.temp/"报告.xlsx"
        report=subprocess.run([sys.executable,str(RUNNER),"report","--case-bundle",str(bundle),"--analysis-json","-","--output",str(output),"--cleanup"],input=json.dumps(analysis,ensure_ascii=False),text=True,capture_output=True)
        self.assertEqual(report.returncode,0,report.stderr)
        self.assertTrue(output.exists()); self.assertFalse(workspace.exists()); self.assertEqual(before,hashlib.sha256(self.evidence.read_bytes()).hexdigest())
        book=load_workbook(output,read_only=True)
        self.assertEqual(book.sheetnames,["汇总","案件事实与请求","证据清单","五层矩阵","矛盾与风险","缺口与补证","法律依据","处理记录"])
        rows=list(book["五层矩阵"].iter_rows(values_only=True)); self.assertEqual(rows[1][4],"部分覆盖")
        book.close()

    def test_invalid_status_rejected_and_workspace_preserved(self):
        process,workspace=self.prepare(); self.assertEqual(process.returncode,0)
        analysis={"五层矩阵":[{"claim":"x","element_and_burden":"x","fact_to_prove":"x","evidence_and_locator":"x","status":"不存在","reason":"x"}]}
        out=self.temp/"bad.xlsx"
        result=subprocess.run([sys.executable,str(RUNNER),"report","--case-bundle",str(workspace/"case-bundle.json"),"--analysis-json","-","--output",str(out)],input=json.dumps(analysis),text=True,capture_output=True)
        self.assertEqual(result.returncode,2); self.assertFalse(out.exists()); self.assertTrue(workspace.exists())

    def test_docx_pdf_xlsx_native_fallback_locations(self):
        docx=self.temp/"解除通知.docx"
        content='''<?xml version="1.0"?><w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body><w:p><w:r><w:t>解除通知内容</w:t></w:r></w:p></w:body></w:document>'''
        types='''<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/></Types>'''
        rels='''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="r" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/></Relationships>'''
        with zipfile.ZipFile(docx,"w") as z: z.writestr("[Content_Types].xml",types); z.writestr("_rels/.rels",rels); z.writestr("word/document.xml",content)
        pdf=self.temp/"工资.pdf"
        pdf.write_bytes(b"%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n2 0 obj<</Type/Pages/Count 1/Kids[3 0 R]>>endobj\n3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 200 200]/Contents 4 0 R/Resources<<>>>>endobj\n4 0 obj<</Length 44>>stream\nBT /F1 12 Tf 10 100 Td (Overtime evidence) Tj ET\nendstream\nendobj\ntrailer<</Root 1 0 R>>\n%%EOF")
        xlsx=self.temp/"考勤.xlsx"; book=Workbook(); book.active.append(["日期","出勤"]); book.active.append(["2026-01-05","加班"]); book.save(xlsx); book.close()
        workspace=self.temp/"binary-workspace"; facts={"summary":"二进制材料回退测试","claims":[]}
        result=subprocess.run([sys.executable,str(RUNNER),"prepare","--facts-json","-","--perspective","neutral","--evidence",str(docx),str(pdf),str(xlsx),"--workspace",str(workspace),"--processing-environment","local","--privacy-confirmed"],input=json.dumps(facts),text=True,capture_output=True)
        self.assertEqual(result.returncode,0,result.stderr)
        bundle=json.loads((workspace/"case-bundle.json").read_text())
        self.assertEqual([x["source_name"] for x in bundle["evidence_files"]],["解除通知.docx","工资.pdf","考勤.xlsx"])
        self.assertTrue(all(x["locators"] for x in bundle["evidence_files"]))

if __name__=="__main__": unittest.main()
